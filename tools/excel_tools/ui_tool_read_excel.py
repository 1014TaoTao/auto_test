# -*- coding:utf-8 -*-

import os

import openpyxl

from common import setting
from tools.public_tool_log import logger


class ExcelOperate:
    # 向excel中写入数据
    """
    @ param Result
    @ param RowNum
    @ param ColNum
    @ param Path
    @ param SheetName
    """

    def __init__(self):
        self.logger = logger(setting.UI_LOG_PATH)

    # 使用openpyxl读取单元格值
    def read_cell(self, excel_path, sheet_name, row_num, cell_num):
        wb = openpyxl.load_workbook(excel_path)
        s1 = wb[sheet_name]
        result = s1.cell(row_num, cell_num).value
        self.logger.info("使用openpyxl读取EXCEL单元格,单元格值为[{0}]".format(result))
        return result

    # 使用openpyxl写excel_path路径的sheet_name中的第row_num行第cell_num列，写入value值
    def write_cell(self, excel_path, sheet_name, row_num, cell_num, value):
        if not os.path.exists(excel_path):
            wb = openpyxl.Workbook()
            sh = wb.create_sheet(sheet_name)
        else:
            wb = openpyxl.load_workbook(excel_path)
            sh = wb[sheet_name]
        sh.cell(row_num, cell_num).value = value
        wb.save(excel_path)
        self.logger.info("使用openpyxl写入EXCEL单元格成功")
        